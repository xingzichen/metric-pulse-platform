"""Run the focused live OMLX regression for three Baijiahao and three World Bank rows.

The Baijiahao input is a previously captured, visually transcribed source snapshot because repeatedly
requesting the historical article can itself trigger a challenge page. World Bank is fetched live through
the production URL adapter. The script never reads the repaired/gold workbook.
"""

from __future__ import annotations

import argparse
import asyncio
import json
import tempfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from metric_pulse.collector import OMLXCollector
from metric_pulse.config import get_settings
from metric_pulse.models import CollectionUnit, DataRecord
from metric_pulse.source_pipeline import (
    SourceDocument,
    normalize_source_url,
    persist_enriched_source_document,
)
from metric_pulse.workbook import find_header_row, make_unique_headers

BAIJIAHAO_ROWS = (8_103, 8_108, 8_113)
WORLD_BANK_ROWS = (6_892, 6_893, 6_894)
EXPECTED = {
    8_103: 16.68,
    8_108: 15.14,
    8_113: 26.81,
    6_892: 2.98956,
    6_893: 3.14297,
    6_894: 3.41788,
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--baijiahao-snapshot", type=Path, required=True)
    return parser.parse_args()


def workbook_rows(path: Path) -> tuple[str, list[str], dict[int, dict[str, Any]]]:
    workbook = load_workbook(path, read_only=False, data_only=False)
    sheet = workbook.worksheets[0]
    header_row = find_header_row(sheet)
    headers = make_unique_headers(
        [sheet.cell(header_row, column).value for column in range(1, sheet.max_column + 1)]
    )
    rows = {
        row_number: {
            header: sheet.cell(row_number, column).value
            for column, header in enumerate(headers, start=1)
        }
        for row_number in (*BAIJIAHAO_ROWS, *WORLD_BANK_ROWS)
    }
    workbook.close()
    return sheet.title, headers, rows


def seed_baijiahao_snapshot(snapshot_path: Path) -> None:
    payload = json.loads(snapshot_path.read_text(encoding="utf-8"))
    normalized_url = normalize_source_url(str(payload["normalized_url"]))
    document = SourceDocument(
        index=1,
        url=str(payload.get("final_url") or normalized_url),
        requested_url=normalized_url,
        normalized_url=normalized_url,
        cache_key=normalized_url,
        title=payload.get("title"),
        media_type=str(payload.get("media_type") or "text/html"),
        text=str(payload["text"]),
        image_table_results=list(payload.get("image_table_results") or []),
        content_hash=payload.get("content_hash"),
    )
    persist_enriched_source_document(document)


def focused_contract(sheet_name: str, source_row: int, row: dict[str, Any]) -> dict[str, Any]:
    descriptors = {
        "index_name": row["index_name"],
        "region": row["region"],
        "statistical_date": row["statistical_date"],
    }
    return {
        "sheet_name": sheet_name,
        "source_row": source_row,
        "descriptors": descriptors,
        "required_matches": list(descriptors),
        "target_fields": ["be_data"],
        "field_roles": {
            "observed": ["be_data"],
            "derived": [],
            "standard_unit": "unit",
            "provenance": "source_url",
        },
        "standard_unit": row["unit"],
        "mode": "focused_same_url_regression",
        "profile": "ai_index_v1",
    }


async def run() -> dict[str, Any]:
    args = parse_args()
    sheet_name, _headers, rows = workbook_rows(args.workbook)
    settings = get_settings()
    network_fetches: list[str] = []

    import metric_pulse.source_pipeline as pipeline

    pipeline._SOURCE_CACHE.clear()
    pipeline._SOURCE_CACHE_LOCKS.clear()
    original_fetch = pipeline.fetch_source_document

    async def counted_fetch(candidate, index, client, validate_url):
        network_fetches.append(candidate.source_url)
        return await original_fetch(candidate, index, client, validate_url)

    pipeline.fetch_source_document = counted_fetch
    seed_baijiahao_snapshot(args.baijiahao_snapshot)
    # Prove that the first row can recover the enriched snapshot from persistent L2, not only process memory.
    pipeline._SOURCE_CACHE.clear()
    collector = OMLXCollector()
    row_reports: list[dict[str, Any]] = []
    try:
        for source_row in (*BAIJIAHAO_ROWS, *WORLD_BANK_ROWS):
            row = rows[source_row]
            record = DataRecord(
                sheet_name=sheet_name,
                source_row=source_row,
                business_key=f"live-regression-{source_row}",
                raw_data=row,
                row_contract=focused_contract(sheet_name, source_row, row),
            )
            result = await collector.collect(record, CollectionUnit(target_fields=["be_data"]))
            calls = result.model_calls[-2:]
            row_reports.append(
                {
                    "source_row": source_row,
                    "region": row["region"],
                    "statistical_date": row["statistical_date"],
                    "expected": EXPECTED[source_row],
                    "actual": result.values.get("be_data"),
                    "passed": result.values.get("be_data") == EXPECTED[source_row],
                    "source_cache_hit": result.acquisition_attempt.get("cache_hit"),
                    "persistent_cache_hit": result.acquisition_attempt.get("persistent_cache_hit"),
                    "model_calls": [
                        {
                            "phase": call["phase"],
                            "finish_reason": call["output_summary"]["provider"].get(
                                "finish_reason"
                            ),
                            "cached_prompt_tokens": call["output_summary"]["provider"].get(
                                "cached_prompt_tokens"
                            ),
                            "usage": call["output_summary"]["provider"].get("usage", {}),
                        }
                        for call in calls
                    ],
                }
            )
    finally:
        pipeline.fetch_source_document = original_fetch
    return {
        "passed": all(item["passed"] for item in row_reports),
        "rows": row_reports,
        "network_fetch_count": len(network_fetches),
        "network_fetch_urls": network_fetches,
        "configured_output_tokens": {
            "default": settings.omlx_max_output_tokens,
            "synthesize": settings.synthesize_max_output_tokens,
            "verify": settings.verify_max_output_tokens,
            "vision_table": settings.vision_table_max_output_tokens,
            "vision_table_retry": settings.vision_table_retry_max_output_tokens,
        },
    }


if __name__ == "__main__":
    with tempfile.TemporaryDirectory(prefix="metric-pulse-live-3x3-") as cache_root:
        get_settings().source_cache_root = Path(cache_root)
        report = asyncio.run(run())
    print(json.dumps(report, ensure_ascii=False, indent=2, default=str))
    raise SystemExit(0 if report["passed"] else 1)
