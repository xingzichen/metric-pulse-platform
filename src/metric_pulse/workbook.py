from __future__ import annotations

import hashlib
import io
import json
import re
from collections import Counter, defaultdict
from copy import copy
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles.numbers import is_date_format
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

MACHINE_HEADER = re.compile(r"^[A-Za-z][A-Za-z0-9_]{1,80}$")
AUDIT_FIELDS = {
    "logic_id",
    "id",
    "created_at",
    "created_time",
    "update_time",
    "updated_at",
    "updater",
    "editor",
    "checker",
    "edit_time",
    "nextcycle_time",
    "data_type",
    "data_status",
}
TARGET_HINTS = {
    "data",
    "be_data",
    "value",
    "result",
    "unit",
    "be_unit",
    "data_unit",
    "source",
    "source_url",
    "source_desc",
    "rank",
    "visits",
    "star",
    "fork",
    "contributors",
    "commits",
}


def json_value(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, bytes):
        return value.hex()
    return value


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def normalized_header(value: Any, index: int) -> str:
    text = str(value).strip() if not is_blank(value) else f"column_{get_column_letter(index)}"
    return text


def find_header_row(ws, *, scan_rows: int = 15) -> int:
    candidates: list[tuple[float, int]] = []
    max_column = min(ws.max_column, 200)
    for row_index in range(1, min(ws.max_row, scan_rows) + 1):
        values = [ws.cell(row_index, col).value for col in range(1, max_column + 1)]
        strings = [str(value).strip() for value in values if not is_blank(value)]
        if not strings:
            continue
        machine = sum(bool(MACHINE_HEADER.fullmatch(value)) for value in strings)
        snake = sum("_" in value for value in strings)
        unique_ratio = len(set(strings)) / len(strings)
        score = machine * 3 + snake + len(strings) * 0.15 + unique_ratio
        if "logic_id" in strings:
            score += 50
        candidates.append((score, row_index))
    if not candidates:
        return 1
    return max(candidates)[1]


def make_unique_headers(values: list[Any]) -> list[str]:
    counts: Counter[str] = Counter()
    headers: list[str] = []
    for index, value in enumerate(values, start=1):
        base = normalized_header(value, index)
        counts[base] += 1
        headers.append(base if counts[base] == 1 else f"{base}__{counts[base]}")
    return headers


def _field_stats(ws, header_row: int, headers: list[str], sample_limit: int = 1000) -> dict[str, dict]:
    row_count = max(0, ws.max_row - header_row)
    sample_count = min(row_count, sample_limit)
    result: dict[str, dict] = {}
    for column, header in enumerate(headers, start=1):
        values = [
            json_value(ws.cell(header_row + offset, column).value) for offset in range(1, sample_count + 1)
        ]
        nonblank = [value for value in values if not is_blank(value)]
        result[header] = {
            "sampled": sample_count,
            "nonblank": len(nonblank),
            "blank_ratio": 1.0 - (len(nonblank) / sample_count) if sample_count else 1.0,
            "unique_ratio": len({json.dumps(v, ensure_ascii=False, default=str) for v in nonblank})
            / len(nonblank)
            if nonblank
            else 0.0,
            "examples": nonblank[:3],
        }
    return result


def _suggest_fields(headers: list[str], stats: dict[str, dict]) -> tuple[list[str], list[str], list[str]]:
    business_keys = [field for field in headers if field in {"logic_id", "id"}]
    if not business_keys:
        business_keys = [
            field
            for field in headers
            if stats[field]["unique_ratio"] > 0.98 and stats[field]["blank_ratio"] < 0.05
        ][:2]

    targets: list[str] = []
    for field in headers:
        lowered = field.lower()
        blank_ratio = stats[field]["blank_ratio"]
        hinted = lowered in TARGET_HINTS or any(
            lowered.endswith(f"_{hint}") for hint in ("value", "data", "unit", "url")
        )
        if field not in AUDIT_FIELDS and hinted and blank_ratio > 0.02:
            targets.append(field)

    if not targets:
        targets = [
            field
            for field in headers
            if field not in AUDIT_FIELDS and 0.2 <= stats[field]["blank_ratio"] < 1.0
        ][:8]

    first_target = min((headers.index(field) for field in targets), default=len(headers))
    descriptors = [
        field
        for field in headers[:first_target]
        if field not in AUDIT_FIELDS and stats[field]["blank_ratio"] < 0.95
    ]
    if not descriptors:
        descriptors = [
            field
            for field in headers
            if field not in set(targets) | AUDIT_FIELDS and stats[field]["blank_ratio"] < 0.5
        ][:10]
    return descriptors, targets, business_keys


def analyze_workbook(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=False, data_only=False)
    sheets: list[dict[str, Any]] = []
    signature_parts: list[str] = []
    for position, ws in enumerate(workbook.worksheets):
        header_row = find_header_row(ws)
        raw_headers = [ws.cell(header_row, col).value for col in range(1, ws.max_column + 1)]
        headers = make_unique_headers(raw_headers)
        display_headers = [
            json_value(ws.cell(max(1, header_row - 1), col).value) for col in range(1, ws.max_column + 1)
        ]
        stats = _field_stats(ws, header_row, headers)
        descriptors, targets, business_keys = _suggest_fields(headers, stats)
        data_rows = 0
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if any(not is_blank(value) for value in row):
                data_rows += 1
        mode = "snapshot_build" if data_rows <= 1 else "row_contract_collect"
        has_machine_headers = sum(bool(MACHINE_HEADER.fullmatch(h)) for h in headers) >= max(
            2, len(headers) // 3
        )
        confidence = 0.96 if has_machine_headers and "logic_id" in headers else 0.72
        needs_confirmation = not targets or confidence < 0.8 or mode == "snapshot_build"
        sheet = {
            "name": ws.title,
            "position": position,
            "header_row": header_row,
            "data_start_row": header_row + 1,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "data_rows": data_rows,
            "headers": headers,
            "display_headers": display_headers,
            "field_stats": stats,
            "descriptor_fields": descriptors,
            "target_fields": targets,
            "business_key_fields": business_keys,
            "mode": mode,
            "confidence": confidence,
            "needs_confirmation": needs_confirmation,
            "merged_ranges": [str(item) for item in ws.merged_cells.ranges],
            "freeze_panes": str(ws.freeze_panes) if ws.freeze_panes else None,
            "hidden": ws.sheet_state != "visible",
        }
        sheets.append(sheet)
        signature_parts.append(json.dumps([ws.title, headers, header_row, ws.max_column], ensure_ascii=False))
    workbook.close()
    return {
        "sheet_count": len(sheets),
        "sheets": sheets,
        "structure_hash": hashlib.sha256("\n".join(signature_parts).encode()).hexdigest(),
        "needs_confirmation": any(sheet["needs_confirmation"] for sheet in sheets),
    }


def render_sheet_preview(path: Path, sheet_name: str, *, max_rows: int = 25, max_cols: int = 16) -> bytes:
    workbook = load_workbook(path, read_only=True, data_only=False)
    ws = workbook[sheet_name]
    rows = min(ws.max_row, max_rows)
    cols = min(ws.max_column, max_cols)
    cell_width = 150
    cell_height = 34
    image = Image.new("RGB", (cell_width * cols, cell_height * rows), "white")
    draw = ImageDraw.Draw(image)
    font = ImageFont.load_default(size=14)
    for row in range(1, rows + 1):
        for col in range(1, cols + 1):
            x0 = (col - 1) * cell_width
            y0 = (row - 1) * cell_height
            fill = "#E8EEF8" if row <= 2 else ("#F8FAFC" if row % 2 else "white")
            draw.rectangle((x0, y0, x0 + cell_width, y0 + cell_height), fill=fill, outline="#CBD5E1")
            coordinate = f"{get_column_letter(col)}{row}"
            value = ws.cell(row, col).value
            text = f"{coordinate}  {'' if value is None else value}"
            if len(text) > 25:
                text = f"{text[:24]}…"
            draw.text((x0 + 5, y0 + 9), text, fill="#0F172A", font=font)
    workbook.close()
    buffer = io.BytesIO()
    image.save(buffer, format="PNG", optimize=True)
    return buffer.getvalue()


def read_rows(
    path: Path,
    *,
    sheet_name: str,
    header_row: int,
    headers: list[str],
) -> list[tuple[int, dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    ws = workbook[sheet_name]
    result: list[tuple[int, dict[str, Any]]] = []
    for row_number, values in enumerate(
        ws.iter_rows(min_row=header_row + 1, max_col=len(headers), values_only=True),
        start=header_row + 1,
    ):
        if not any(not is_blank(value) for value in values):
            continue
        result.append(
            (
                row_number,
                {header: json_value(values[index]) for index, header in enumerate(headers)},
            )
        )
    workbook.close()
    return result


def export_reviewed_workbook(
    source: Path,
    destination: Path,
    updates: list[tuple[str, int, dict[str, Any]]],
    *,
    column_definitions: dict[str, list[tuple[str, Any, Any]]] | None = None,
    sheet_dimensions: dict[str, tuple[int, int]] | None = None,
) -> None:
    workbook = load_workbook(source, read_only=False, data_only=False)
    column_definitions = column_definitions or {}
    sheet_dimensions = sheet_dimensions or {}
    grouped: dict[str, list[tuple[int, dict[str, Any]]]] = defaultdict(list)
    for sheet_name, row_number, values in updates:
        grouped[sheet_name].append((row_number, values))
    for sheet_name, sheet_updates in grouped.items():
        ws = workbook[sheet_name]
        header_row = find_header_row(ws)
        columns = {
            header: index
            for index, header in enumerate(
                make_unique_headers([ws.cell(header_row, col).value for col in range(1, ws.max_column + 1)]),
                start=1,
            )
        }
        for field, display_header, machine_header in column_definitions.get(sheet_name, []):
            if field in columns:
                continue
            column = ws.max_column + 1
            previous = max(1, column - 1)
            for row in (max(1, header_row - 1), header_row, header_row + 1):
                ws.cell(row, column)._style = copy(ws.cell(row, previous)._style)
            ws.cell(max(1, header_row - 1), column).value = display_header
            ws.cell(header_row, column).value = machine_header
            columns[field] = column
        initial_max_row = ws.max_row
        template_row = min(header_row + 1, initial_max_row)
        for row_number, values in sheet_updates:
            if row_number > initial_max_row:
                for column in range(1, ws.max_column + 1):
                    source_cell = ws.cell(template_row, column)
                    target_cell = ws.cell(row_number, column)
                    target_cell._style = copy(source_cell._style)
                    if source_cell.has_style:
                        target_cell.number_format = source_cell.number_format
            for field, value in values.items():
                if field not in columns:
                    raise ValueError(f"Unknown field {field!r} in sheet {sheet_name!r}")
                cell = ws.cell(row_number, columns[field])
                original_style = copy(cell._style)
                cell.value = value
                cell._style = original_style
                if isinstance(value, (date, datetime)) and not is_date_format(cell.number_format):
                    cell.number_format = "yyyy-mm-dd"
    for sheet_name, (target_rows, target_columns) in sheet_dimensions.items():
        ws = workbook[sheet_name]
        if target_rows > ws.max_row or target_columns > ws.max_column:
            source_row = min(find_header_row(ws) + 1, ws.max_row)
            source_column = min(ws.max_column, max(1, target_columns))
            anchor = ws.cell(target_rows, target_columns)
            anchor._style = copy(ws.cell(source_row, source_column)._style)
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    workbook.close()
