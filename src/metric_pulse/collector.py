from __future__ import annotations

import asyncio
import ipaddress
import json
import socket
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Protocol
from urllib.parse import urlparse

import httpx
from openpyxl import load_workbook

from .config import get_settings
from .models import CollectionUnit, DataRecord
from .omlx import OMLXClient
from .workbook import find_header_row, make_unique_headers


@dataclass(slots=True)
class EvidenceItem:
    source_url: str | None = None
    title: str | None = None
    locator: str | None = None
    excerpt: str | None = None
    metadata: dict[str, Any] = field(default_factory=dict)


@dataclass(slots=True)
class CollectionResult:
    values: dict[str, Any]
    evidence: list[EvidenceItem] = field(default_factory=list)
    validation: dict[str, Any] = field(default_factory=lambda: {"valid": True, "errors": []})
    model: str | None = None


class Collector(Protocol):
    async def collect(self, record: DataRecord, unit: CollectionUnit) -> CollectionResult: ...


async def fetch_public_text(url: str, *, max_bytes: int = 2_000_000) -> str:
    await validate_public_url(url)
    async with httpx.AsyncClient(follow_redirects=True, timeout=20) as client:
        response = await client.get(url, headers={"User-Agent": "MetricPulse/1.0"})
        response.raise_for_status()
        await validate_public_url(str(response.url))
        content = response.content[:max_bytes]
        return content.decode(response.encoding or "utf-8", errors="replace")


async def validate_public_url(url: str) -> None:
    parsed = urlparse(url)
    if parsed.scheme not in {"http", "https"} or not parsed.hostname:
        raise ValueError("Only public HTTP(S) URLs are allowed")
    try:
        addresses = await asyncio.to_thread(
            socket.getaddrinfo,
            parsed.hostname,
            parsed.port or (443 if parsed.scheme == "https" else 80),
        )
    except socket.gaierror as exc:
        raise ValueError("Evidence host cannot be resolved") from exc
    for address in addresses:
        ip = ipaddress.ip_address(address[4][0])
        if not ip.is_global:
            raise ValueError("Private, loopback, and reserved evidence addresses are not allowed")


async def discover_source(query: str) -> tuple[str | None, str | None]:
    """Discover one public source through an optional SearXNG-compatible endpoint."""
    search_url = get_settings().search_url
    if not search_url or not query.strip():
        return None, None
    async with httpx.AsyncClient(timeout=20) as client:
        response = await client.get(search_url, params={"q": query, "format": "json"})
        response.raise_for_status()
    for item in response.json().get("results", []):
        candidate = item.get("url")
        if not isinstance(candidate, str):
            continue
        try:
            await validate_public_url(candidate)
        except ValueError:
            continue
        return candidate, item.get("title")
    return None, None


class OMLXCollector:
    def __init__(self, client: OMLXClient | None = None) -> None:
        self.client = client or OMLXClient()

    async def collect(self, record: DataRecord, unit: CollectionUnit) -> CollectionResult:
        source_url = next(
            (
                value
                for key, value in record.raw_data.items()
                if key in {"source_url", "url", "link"} and isinstance(value, str) and value
            ),
            None,
        )
        source_title = None
        if not source_url:
            descriptors = record.row_contract.get("descriptors", {})
            query = " ".join(str(value) for value in descriptors.values() if value not in (None, ""))
            try:
                source_url, source_title = await discover_source(query)
            except (httpx.HTTPError, ValueError, KeyError):
                source_url = None
        evidence_text = ""
        if source_url:
            try:
                evidence_text = (await fetch_public_text(source_url))[:30_000]
            except httpx.HTTPError, ValueError:
                evidence_text = ""
        prompt = {
            "row_contract": record.row_contract,
            "raw_row": record.raw_data,
            "target_fields": unit.target_fields,
            "evidence_text": evidence_text,
            "requirements": {
                "return": {"values": {field: "value or null" for field in unit.target_fields}},
                "do_not_invent": True,
            },
        }
        response = await self.client.generate_json(
            system=(
                "Extract only the requested target fields. Respect every RowContract descriptor. "
                "Return JSON with a values object and a short evidence_excerpt. Use null when unsupported."
            ),
            prompt=json.dumps(prompt, ensure_ascii=False, default=str),
        )
        raw_values = response.get("values", {})
        values = {field: raw_values.get(field) for field in unit.target_fields}
        evidence = []
        if source_url or response.get("evidence_excerpt"):
            evidence.append(
                EvidenceItem(
                    source_url=source_url,
                    title=source_title,
                    excerpt=response.get("evidence_excerpt"),
                    metadata={"provider": "omlx"},
                )
            )
        errors = [field for field, value in values.items() if value is None]
        return CollectionResult(
            values=values,
            evidence=evidence,
            validation={"valid": not errors, "missing_fields": errors},
            model=get_settings().omlx_model,
        )


class GoldWorkbookCollector:
    """Deterministic acceptance-test adapter; never selected in production by default."""

    def __init__(self, path: Path) -> None:
        self.workbook = load_workbook(path, read_only=True, data_only=False)
        self.columns: dict[str, dict[str, int]] = {}
        for ws in self.workbook.worksheets:
            header_row = find_header_row(ws)
            headers = make_unique_headers(
                [ws.cell(header_row, column).value for column in range(1, ws.max_column + 1)]
            )
            self.columns[ws.title] = {header: index for index, header in enumerate(headers, start=1)}

    async def collect(self, record: DataRecord, unit: CollectionUnit) -> CollectionResult:
        ws = self.workbook[record.sheet_name]
        columns = self.columns[record.sheet_name]
        values = {field: ws.cell(record.source_row, columns[field]).value for field in unit.target_fields}
        return CollectionResult(
            values=values,
            evidence=[
                EvidenceItem(
                    title="Acceptance gold workbook",
                    locator=f"{record.sheet_name}!row:{record.source_row}",
                    metadata={"provider": "gold-fixture"},
                )
            ],
            validation={"valid": True, "fixture": True},
            model="gold-fixture",
        )


def configured_collector() -> Collector:
    settings = get_settings()
    if settings.collector_mode == "gold":
        if not settings.gold_workbook_path:
            raise RuntimeError("MP_GOLD_WORKBOOK_PATH is required for gold collector mode")
        return GoldWorkbookCollector(settings.gold_workbook_path)
    return OMLXCollector()
