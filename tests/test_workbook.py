from __future__ import annotations

from openpyxl import load_workbook

from metric_pulse.dataset_profiles import (
    AI_ALGORITHM_COLLECTION_TARGET_FIELDS,
    TOP_LIST_AI_TARGET_FIELDS,
)
from metric_pulse.workbook import analyze_workbook, render_sheet_preview, top_list_ai_batch_state
from tests.conftest import EMPTY_WORKBOOK


def test_real_fixture_analysis_recognizes_all_sheets() -> None:
    analysis = analyze_workbook(EMPTY_WORKBOOK)
    assert analysis["sheet_count"] == 11
    assert len(analysis["structure_hash"]) == 64
    assert all(sheet["header_row"] == 2 for sheet in analysis["sheets"])
    assert all("logic_id" in sheet["headers"] for sheet in analysis["sheets"])
    assert any(sheet["data_rows"] > 10_000 for sheet in analysis["sheets"])
    algorithm = next(
        sheet for sheet in analysis["sheets"] if "ai_algorithm_collectio" in sheet["name"].lower()
    )
    assert algorithm["mode"] == "monthly_top10_append"
    assert algorithm["target_fields"] == list(AI_ALGORITHM_COLLECTION_TARGET_FIELDS)
    ai_index = next(sheet for sheet in analysis["sheets"] if "ai_index" in sheet["name"].lower())
    assert ai_index["target_fields"] == ["be_data", "be_unit", "data", "source_url"]
    assert "district" in ai_index["descriptor_fields"]
    assert "other_region" in ai_index["descriptor_fields"]
    top_list = next(sheet for sheet in analysis["sheets"] if "top_list_ai" in sheet["name"].lower())
    assert top_list["mode"] == "annual_top50_append"
    assert top_list["target_fields"] == list(TOP_LIST_AI_TARGET_FIELDS)
    excluded = [sheet for sheet in analysis["sheets"] if sheet["excluded"]]
    assert {sheet["exclusion_reason"]["sheet_id"] for sheet in excluded} == {
        "ai_news",
        "gpu_chip_performance",
        "ai_model_permission",
        "aigc_reg_i",
        "ai_person",
        "ai_computing_power",
    }
    assert all(sheet["mode"] == "excluded" for sheet in excluded)
    assert all(sheet["target_fields"] == [] for sheet in excluded)
    assert all(sheet["descriptor_fields"] == [] for sheet in excluded)
    assert all(sheet["business_key_fields"] == [] for sheet in excluded)
    assert all(sheet["needs_confirmation"] is False for sheet in excluded)


def test_real_fixture_top_list_requires_new_official_batch() -> None:
    analysis = analyze_workbook(EMPTY_WORKBOOK)
    top_list = next(sheet for sheet in analysis["sheets"] if "top_list_ai" in sheet["name"].lower())

    state = top_list_ai_batch_state(
        EMPTY_WORKBOOK,
        sheet_name=top_list["name"],
        header_row=top_list["header_row"],
        headers=top_list["headers"],
        rank_year=2026,
    )

    assert state["idempotent"] is False
    assert state["superseded_rows"] == list(range(153, 203))


def test_real_fixture_preview_is_png() -> None:
    workbook = load_workbook(EMPTY_WORKBOOK, read_only=True)
    sheet_name = workbook.sheetnames[0]
    workbook.close()
    preview = render_sheet_preview(EMPTY_WORKBOOK, sheet_name, max_rows=6, max_cols=8)
    assert preview.startswith(b"\x89PNG\r\n\x1a\n")
    assert len(preview) > 1_000
