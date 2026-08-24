from __future__ import annotations

import asyncio
import io
from copy import copy
from datetime import datetime

from openpyxl import load_workbook

from metric_pulse import main as main_module
from metric_pulse.dataset_profiles import (
    FORBES_AI50_SOURCE_URL,
    GITHUB_TOP_REPOSITORIES_SOURCE_URL,
    is_excluded_sheet,
)
from tests.conftest import EMPTY_WORKBOOK, TEST_ROOT


def test_upload_collect_review_and_export_real_fixture(client) -> None:
    with EMPTY_WORKBOOK.open("rb") as handle:
        response = client.post(
            "/api/v1/files",
            files={
                "upload": (
                    EMPTY_WORKBOOK.name,
                    handle,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    assert response.status_code == 201, response.text
    file_payload = response.json()
    assert file_payload["analysis"]["sheet_count"] == 11
    target_sheet = next(
        sheet for sheet in file_payload["analysis"]["sheets"] if sheet["name"].startswith("人工智能算法收藏")
    )

    response = client.post(
        "/api/v1/tasks",
        json={
            "file_id": file_payload["id"],
            "name": "real-fixture-snapshot",
            "datasets": [
                {
                    "sheet_name": target_sheet["name"],
                    "descriptor_fields": ["collect_date", "rank"],
                    "target_fields": ["name", "star", "star_unit", "source_url"],
                    "business_key_fields": ["logic_id"],
                    "mode": "snapshot_build",
                }
            ],
        },
    )
    assert response.status_code == 201, response.text
    task = response.json()
    # 月度增量 Profile 不读取外部对照文件，也不把已有历史样例当作本次快照。
    assert task["stats"]["total"] == 10

    response = client.post(
        f"/api/v1/tasks/{task['id']}/start",
        json={"expected_version": task["version"]},
    )
    assert response.status_code == 202, response.text
    task = client.get(f"/api/v1/tasks/{task['id']}").json()
    assert task["status"] == "SUCCEEDED"
    assert task["stats"]["succeeded"] == task["stats"]["total"]

    units = []
    offset = 0
    while True:
        page = client.get(
            f"/api/v1/tasks/{task['id']}/review-queue",
            params={"offset": offset, "limit": 200},
        ).json()
        units.extend(page["items"])
        offset += len(page["items"])
        if offset >= page["total"]:
            break
    for start in range(0, len(units), 200):
        preview = client.post(
            f"/api/v1/tasks/{task['id']}/reviews/bulk/preview",
            json={"unit_ids": [item["id"] for item in units[start : start + 200]]},
        )
        assert preview.status_code == 200, preview.text
        assert preview.json()["eligible"] == len(units[start : start + 200])
        response = client.post(
            f"/api/v1/tasks/{task['id']}/reviews/bulk/commit",
            json={"preview_token": preview.json()["previewToken"]},
        )
        assert response.status_code == 200, response.text

    readiness = client.get(f"/api/v1/tasks/{task['id']}/export-readiness").json()
    assert readiness == {"ready": True, "blockers": [], "counts": {"APPROVED": len(units)}}
    response = client.post(f"/api/v1/tasks/{task['id']}/exports")
    assert response.status_code == 201, response.text
    export = response.json()
    assert export["status"] == "READY"
    download = client.get(f"/api/v1/exports/{export['id']}/download")
    assert download.status_code == 200

    source = load_workbook(EMPTY_WORKBOOK, read_only=False, data_only=False)
    actual_path = TEST_ROOT / "api-acceptance-output.xlsx"
    actual_path.parent.mkdir(parents=True, exist_ok=True)
    actual_path.write_bytes(download.content)
    actual = load_workbook(actual_path, read_only=False, data_only=False)
    # 不由平台处理的工作表仍跟随原工作簿导出，但任何单元格和结构都不能被本任务改写。
    for sheet_name in source.sheetnames:
        if not is_excluded_sheet(sheet_name):
            continue
        source_sheet = source[sheet_name]
        actual_sheet = actual[sheet_name]
        assert actual_sheet.max_row == source_sheet.max_row
        assert actual_sheet.max_column == source_sheet.max_column
        assert list(actual_sheet.merged_cells.ranges) == list(source_sheet.merged_cells.ranges)
        assert actual_sheet.freeze_panes == source_sheet.freeze_panes
        assert actual_sheet.sheet_state == source_sheet.sheet_state
        assert {
            key: (item.height, item.hidden, item.outline_level, item.collapsed)
            for key, item in actual_sheet.row_dimensions.items()
        } == {
            key: (item.height, item.hidden, item.outline_level, item.collapsed)
            for key, item in source_sheet.row_dimensions.items()
        }
        assert {
            key: (
                item.width,
                item.hidden,
                    item.bestFit,
                item.outline_level,
                item.collapsed,
                item.min,
                item.max,
            )
            for key, item in actual_sheet.column_dimensions.items()
        } == {
            key: (
                item.width,
                item.hidden,
                    item.bestFit,
                item.outline_level,
                item.collapsed,
                item.min,
                item.max,
            )
            for key, item in source_sheet.column_dimensions.items()
        }
        for source_row, actual_row in zip(
            source_sheet.iter_rows(),
            actual_sheet.iter_rows(),
            strict=True,
        ):
            assert [None if cell.value == "" else cell.value for cell in actual_row] == [
                None if cell.value == "" else cell.value for cell in source_row
            ]
            assert [
                (
                    copy(cell.font),
                    copy(cell.fill),
                    copy(cell.border),
                    copy(cell.alignment),
                    cell.number_format,
                    copy(cell.protection),
                )
                for cell in actual_row
            ] == [
                (
                    copy(cell.font),
                    copy(cell.fill),
                    copy(cell.border),
                    copy(cell.alignment),
                    cell.number_format,
                    copy(cell.protection),
                )
                for cell in source_row
            ]
    # 历史样例保持原样，新快照使用其后的十个预格式化空白行。
    for column in range(1, source[target_sheet["name"]].max_column + 1):
        assert (
            actual[target_sheet["name"]].cell(3, column).value
            == source[target_sheet["name"]].cell(3, column).value
        )
    headers = {
        actual[target_sheet["name"]].cell(2, column).value: column
        for column in range(1, actual[target_sheet["name"]].max_column + 1)
    }
    snapshot_times = set()
    for rank, row in enumerate(range(4, 14), start=1):
        assert actual[target_sheet["name"]].cell(row, headers["rank"]).value == rank
        assert actual[target_sheet["name"]].cell(row, headers["name"]).value == f"fixture/repository-{rank}"
        assert actual[target_sheet["name"]].cell(row, headers["star"]).value == 1_001 - rank
        assert actual[target_sheet["name"]].cell(row, headers["star_unit"]).value == "k"
        assert (
            actual[target_sheet["name"]].cell(row, headers["source_url"]).value
            == GITHUB_TOP_REPOSITORIES_SOURCE_URL
        )
        assert actual[target_sheet["name"]].cell(row, headers["source_department"]).value == "Github"
        assert actual[target_sheet["name"]].cell(row, headers["update_frequency"]).value == "month"
        assert actual[target_sheet["name"]].cell(row, headers["data_type"]).value == "采集"
        assert actual[target_sheet["name"]].cell(row, headers["data_status"]).value == "新增"
        assert actual[target_sheet["name"]].cell(row, headers["update_time"]).value is None
        assert actual[target_sheet["name"]].cell(row, headers["created_time"]).value is None
        timestamps = {
            actual[target_sheet["name"]].cell(row, headers[field]).value
            for field in ("collect_date", "datasource_date", "collection_date")
        }
        assert len(timestamps) == 1
        assert all(isinstance(value, datetime) for value in timestamps)
        snapshot_times.update(timestamps)
    assert len(snapshot_times) == 1
    actual.close()
    source.close()


def test_forbes_ai50_annual_batch_switch_and_idempotent_replan(client) -> None:
    with EMPTY_WORKBOOK.open("rb") as handle:
        response = client.post(
            "/api/v1/files",
            files={
                "upload": (
                    EMPTY_WORKBOOK.name,
                    handle,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    assert response.status_code == 201, response.text
    file_payload = response.json()
    target_sheet = next(
        sheet for sheet in file_payload["analysis"]["sheets"] if "top_list_ai" in sheet["name"].lower()
    )

    response = client.post(
        "/api/v1/tasks",
        json={
            "file_id": file_payload["id"],
            "name": "forbes-ai50-annual-fixture",
            "datasets": [
                {
                    "sheet_name": target_sheet["name"],
                    "descriptor_fields": target_sheet["descriptor_fields"],
                    "target_fields": target_sheet["target_fields"],
                    "business_key_fields": target_sheet["business_key_fields"],
                    "mode": target_sheet["mode"],
                }
            ],
        },
    )
    assert response.status_code == 201, response.text
    task = response.json()
    assert task["stats"]["total"] == 50

    response = client.post(
        f"/api/v1/tasks/{task['id']}/start",
        json={"expected_version": task["version"]},
    )
    assert response.status_code == 202, response.text
    task = client.get(f"/api/v1/tasks/{task['id']}").json()
    assert task["status"] == "SUCCEEDED"
    assert task["stats"]["succeeded"] == 50

    queue = client.get(
        f"/api/v1/tasks/{task['id']}/review-queue",
        params={"offset": 0, "limit": 100},
    ).json()
    assert queue["total"] == 50
    source_rows = sorted(
        client.get(f"/api/v1/review-units/{item['id']}").json()["record"]["sourceRow"]
        for item in queue["items"]
    )
    assert source_rows == list(range(203, 253))
    preview = client.post(
        f"/api/v1/tasks/{task['id']}/reviews/bulk/preview",
        json={"unit_ids": [item["id"] for item in queue["items"]]},
    )
    assert preview.status_code == 200, preview.text
    response = client.post(
        f"/api/v1/tasks/{task['id']}/reviews/bulk/commit",
        json={"preview_token": preview.json()["previewToken"]},
    )
    assert response.status_code == 200, response.text

    readiness = client.get(f"/api/v1/tasks/{task['id']}/export-readiness").json()
    assert readiness["ready"] is True
    response = client.post(f"/api/v1/tasks/{task['id']}/exports")
    assert response.status_code == 201, response.text
    export_id = response.json()["id"]
    downloaded = client.get(f"/api/v1/exports/{export_id}/download")
    assert downloaded.status_code == 200, downloaded.text

    workbook = load_workbook(io.BytesIO(downloaded.content), data_only=False)
    sheet = workbook[target_sheet["name"]]
    headers = {sheet.cell(2, column).value: column for column in range(1, sheet.max_column + 1)}
    assert all(sheet.cell(row, headers["data_status"]).value == "删除" for row in range(153, 203))
    assert all(sheet.cell(row, headers["data_status"]).value == "新增" for row in range(203, 253))
    assert all(
        sheet.cell(row, headers["source_url"]).value == FORBES_AI50_SOURCE_URL
        for row in range(203, 253)
    )
    assert [sheet.cell(row, headers["company_name"]).value for row in range(203, 253)] == [
        f"Fixture AI Company {position:02d}" for position in range(1, 51)
    ]
    workbook.close()

    response = client.post(
        "/api/v1/files",
        files={
            "upload": (
                "forbes-ai50-export.xlsx",
                downloaded.content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 201, response.text
    exported_file = response.json()
    exported_sheet = next(
        sheet for sheet in exported_file["analysis"]["sheets"] if "top_list_ai" in sheet["name"].lower()
    )
    response = client.post(
        "/api/v1/tasks",
        json={
            "file_id": exported_file["id"],
            "name": "forbes-ai50-idempotency",
            "datasets": [
                {
                    "sheet_name": exported_sheet["name"],
                    "descriptor_fields": exported_sheet["descriptor_fields"],
                    "target_fields": exported_sheet["target_fields"],
                    "business_key_fields": exported_sheet["business_key_fields"],
                    "mode": exported_sheet["mode"],
                }
            ],
        },
    )
    assert response.status_code == 201, response.text
    assert response.json()["stats"]["total"] == 0


def test_default_task_excludes_out_of_scope_sheets_and_rejects_manual_override(client) -> None:
    with EMPTY_WORKBOOK.open("rb") as handle:
        response = client.post(
            "/api/v1/files",
            files={
                "upload": (
                    EMPTY_WORKBOOK.name,
                    handle,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    assert response.status_code == 201, response.text
    file_payload = response.json()
    excluded = [sheet for sheet in file_payload["analysis"]["sheets"] if sheet["excluded"]]
    assert len(excluded) == 6

    response = client.post(
        "/api/v1/tasks",
        json={
            "file_id": file_payload["id"],
            "name": "all-supported-sheets",
            "datasets": [],
        },
    )
    assert response.status_code == 201, response.text
    task = response.json()
    assert task["stats"]["total"] == 10_775
    planned_names = {dataset["sheet_name"] for dataset in task["config"]["datasets"]}
    assert planned_names.isdisjoint({sheet["name"] for sheet in excluded})

    excluded_sheet = excluded[0]
    response = client.post(
        "/api/v1/tasks",
        json={
            "file_id": file_payload["id"],
            "name": "forbidden-manual-override",
            "datasets": [
                {
                    "sheet_name": excluded_sheet["name"],
                    "descriptor_fields": [],
                    "target_fields": [excluded_sheet["headers"][0]],
                    "business_key_fields": [],
                    "mode": "row_contract_collect",
                }
            ],
        },
    )
    assert response.status_code == 422, response.text
    assert "不由本平台处理" in response.json()["detail"]


def test_vision_preflight_failure_skips_out_of_scope_sheets(client, monkeypatch) -> None:
    async def unavailable(_self):
        raise RuntimeError("provider rejected credentials")

    monkeypatch.setattr(main_module.settings, "vision_analysis_enabled", False)
    with EMPTY_WORKBOOK.open("rb") as handle:
        response = client.post(
            "/api/v1/files",
            files={
                "upload": (
                    EMPTY_WORKBOOK.name,
                    handle,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    assert response.status_code == 201, response.text
    file_id = response.json()["id"]

    monkeypatch.setattr(main_module.settings, "vision_analysis_enabled", True)
    monkeypatch.setattr(main_module.OMLXClient, "health", unavailable)
    asyncio.run(main_module.vision_recognize_file(file_id))

    detail = client.get(f"/api/v1/files/{file_id}")
    assert detail.status_code == 200, detail.text
    sheets = detail.json()["analysis"]["sheets"]
    assert sheets
    excluded = [sheet for sheet in sheets if sheet["excluded"]]
    locked_profiles = [
        sheet
        for sheet in sheets
        if not sheet["excluded"]
        and sheet["vision"].get("reason", {}).get("code") == "PROFILE_CONTRACT_LOCKED"
    ]
    processable = [sheet for sheet in sheets if not sheet["vision"].get("skipped")]
    assert len(excluded) == 6
    assert len(locked_profiles) == 3
    assert all(sheet["vision"]["skipped"] is True for sheet in excluded)
    assert all(sheet["vision"]["roleOverrideLocked"] is True for sheet in excluded)
    assert all(sheet["needs_confirmation"] is False for sheet in excluded)
    assert all(sheet["vision"]["valid"] is False for sheet in processable)
    assert all("OMLX preflight failed" in sheet["vision"]["error"] for sheet in processable)
