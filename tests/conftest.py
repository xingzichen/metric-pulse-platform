from __future__ import annotations

import hashlib
import os
import shutil
from collections.abc import Generator
from pathlib import Path

import pytest
from openpyxl import load_workbook

TEST_ROOT = Path("/private/tmp/metric-pulse-platform-tests")
WORKSPACE_ROOT = Path(__file__).resolve().parents[2]
EMPTY_WORKBOOK = WORKSPACE_ROOT / "人工智能-空表--250224-V1.xlsx"
GOLD_WORKBOOK = WORKSPACE_ROOT / "人工智能-修复表--250224-V1.xlsx"

os.environ.update(
    {
        "MP_ENV": "test",
        "MP_DATABASE_URL": f"sqlite:///{TEST_ROOT / 'test.db'}",
        "MP_OBJECT_ROOT": str(TEST_ROOT / "objects"),
        "MP_EXPORT_ROOT": str(TEST_ROOT / "exports"),
        "MP_BOOTSTRAP_USERNAME": "admin",
        "MP_BOOTSTRAP_PASSWORD": "test-password",
        "MP_VISION_ANALYSIS_ENABLED": "false",
        "MP_EAGER_TASKS": "true",
    }
)

from fastapi.testclient import TestClient  # noqa: E402

from metric_pulse import main as main_module  # noqa: E402
from metric_pulse import processor as processor_module  # noqa: E402
from metric_pulse.collector import CollectionResult, EvidenceItem  # noqa: E402
from metric_pulse.dataset_profiles import (  # noqa: E402
    AI_ALGORITHM_COLLECTION_PROFILE,
    FORBES_AI50_SOURCE_URL,
    GITHUB_TOP_REPOSITORIES_SOURCE_URL,
    TOP_LIST_AI_PROFILE,
)
from metric_pulse.db import Base, SessionLocal, engine  # noqa: E402
from metric_pulse.main import app  # noqa: E402
from metric_pulse.security import bootstrap_admin  # noqa: E402
from metric_pulse.workbook import find_header_row, make_unique_headers  # noqa: E402


class FixtureWorkbookCollector:
    """Test-only pipeline double. Production code never imports the expected workbook."""

    def __init__(self) -> None:
        self.workbook = load_workbook(GOLD_WORKBOOK, read_only=True, data_only=False)
        self.columns: dict[str, dict[str, int]] = {}
        for sheet in self.workbook.worksheets:
            header_row = find_header_row(sheet)
            headers = make_unique_headers(
                [sheet.cell(header_row, column).value for column in range(1, sheet.max_column + 1)]
            )
            self.columns[sheet.title] = {
                header: index for index, header in enumerate(headers, start=1)
            }

    async def collect(self, record, unit) -> CollectionResult:
        if record.row_contract.get("profile") == AI_ALGORITHM_COLLECTION_PROFILE:
            rank = int(record.row_contract["rank"])
            name = f"fixture/repository-{rank}"
            snapshot_at = record.row_contract["snapshot_at"]
            values = dict.fromkeys(unit.target_fields)
            values.update(record.row_contract["fixed_values"])
            values.update(
                {
                    "logic_id": hashlib.sha256(f"{name}\n{snapshot_at}".encode()).hexdigest(),
                    "name": name,
                    "star": 1_001 - rank,
                    "source_url": GITHUB_TOP_REPOSITORIES_SOURCE_URL,
                }
            )
            return CollectionResult(
                values=values,
                evidence=[
                    EvidenceItem(
                        source_url=GITHUB_TOP_REPOSITORIES_SOURCE_URL,
                        title="GitHub repository search fixture",
                        metadata={"provider": "test-source-fixture", "selected": True},
                    )
                ],
                validation={"valid": True, "source_fixture": True},
                model="test-fixture",
            )
        if record.row_contract.get("profile") == TOP_LIST_AI_PROFILE:
            position = int(record.row_contract["list_position"])
            rank_year = int(record.row_contract["rank_year"])
            name = f"Fixture AI Company {position:02d}"
            normalized_name = " ".join(name.casefold().split())
            values = dict.fromkeys(unit.target_fields)
            values.update(record.row_contract["fixed_values"])
            values.update(
                {
                    "logic_id": hashlib.sha256(
                        f"{rank_year}\n{normalized_name}".encode()
                    ).hexdigest(),
                    "rank_year": rank_year,
                    "company_name": name,
                    "headquarter_location": f"测试总部 {position}",
                    "CEO": f"Test CEO {position}",
                    "financing_amount": position / 10,
                    "financing_amount_unit": "亿美元",
                    "establish_date": 2000 + position % 20,
                    "source": "福布斯",
                    "source_url": FORBES_AI50_SOURCE_URL,
                    "update_frequency": "year",
                    "datasource_date": f"{rank_year}-04-16T06:30:00-04:00",
                    "collection_date": record.row_contract["snapshot_at"],
                    "data_type": "采集",
                    "data_status": "新增",
                }
            )
            return CollectionResult(
                values=values,
                evidence=[
                    EvidenceItem(
                        source_url=FORBES_AI50_SOURCE_URL,
                        title="Forbes official AI 50 fixture",
                        metadata={"provider": "test-source-fixture", "selected": True},
                    )
                ],
                validation={
                    "valid": True,
                    "source_fixture": True,
                    "dataset_profile": TOP_LIST_AI_PROFILE,
                },
                model="test-fixture",
            )
        sheet = self.workbook[record.sheet_name]
        columns = self.columns[record.sheet_name]
        values = {
            field: sheet.cell(record.source_row, columns[field]).value
            for field in unit.target_fields
        }
        return CollectionResult(
            values=values,
            evidence=[
                EvidenceItem(
                    title="External fixture adapter",
                    locator=f"{record.sheet_name}!row:{record.source_row}",
                    metadata={"provider": "test-fixture"},
                )
            ],
            validation={"valid": True, "fixture": True},
            model="test-fixture",
        )


@pytest.fixture()
def client(monkeypatch: pytest.MonkeyPatch) -> Generator[TestClient]:
    async def model_ready(_self) -> dict[str, object]:
        return {"ok": True, "models": []}

    monkeypatch.setattr(main_module.OMLXClient, "health", model_ready)
    monkeypatch.setattr(processor_module, "configured_collector", FixtureWorkbookCollector)
    engine.dispose()
    if TEST_ROOT.exists():
        shutil.rmtree(TEST_ROOT)
    TEST_ROOT.mkdir(parents=True)
    Base.metadata.drop_all(engine)
    Base.metadata.create_all(engine)
    with SessionLocal() as db:
        bootstrap_admin(db)
    with TestClient(app) as test_client:
        response = test_client.post(
            "/api/v1/auth/login",
            json={"username": "admin", "password": "test-password"},
        )
        assert response.status_code == 200, response.text
        yield test_client
