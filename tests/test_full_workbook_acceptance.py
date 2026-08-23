from __future__ import annotations

from collections import defaultdict
from itertools import zip_longest

import pytest
from openpyxl import load_workbook

from metric_pulse.workbook import export_reviewed_workbook, find_header_row, make_unique_headers
from tests.conftest import EMPTY_WORKBOOK, GOLD_WORKBOOK, TEST_ROOT


def normalize_blank(value):
    return None if value == "" else value


@pytest.mark.acceptance
def test_full_fixture_can_be_rebuilt_to_gold_values() -> None:
    source = load_workbook(EMPTY_WORKBOOK, read_only=True, data_only=False)
    gold = load_workbook(GOLD_WORKBOOK, read_only=True, data_only=False)
    updates: list[tuple[str, int, dict]] = []
    column_definitions: dict[str, list[tuple[str, object, object]]] = {}
    sheet_dimensions: dict[str, tuple[int, int]] = {}
    expected_changes = 0
    for sheet_name in source.sheetnames:
        source_ws = source[sheet_name]
        gold_ws = gold[sheet_name]
        header_row = find_header_row(source_ws)
        source_headers = make_unique_headers(
            list(next(source_ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)))
        )
        raw_gold_headers = list(
            next(gold_ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
        )
        headers = make_unique_headers(raw_gold_headers)
        display_headers = list(
            next(gold_ws.iter_rows(min_row=header_row - 1, max_row=header_row - 1, values_only=True))
        )
        column_definitions[sheet_name] = [
            (
                field,
                display_headers[index] if index < len(display_headers) else field,
                raw_gold_headers[index],
            )
            for index, field in enumerate(headers)
            if field not in source_headers
        ]
        sheet_dimensions[sheet_name] = (gold_ws.max_row, gold_ws.max_column)
        source_rows = source_ws.iter_rows(min_row=header_row + 1, max_col=len(headers), values_only=True)
        gold_rows = gold_ws.iter_rows(min_row=header_row + 1, max_col=len(headers), values_only=True)
        for row_number, (source_row, gold_row) in enumerate(
            zip_longest(source_rows, gold_rows, fillvalue=()), start=header_row + 1
        ):
            changed: dict = {}
            for column, header in enumerate(headers):
                source_value = source_row[column] if column < len(source_row) else None
                gold_value = gold_row[column] if column < len(gold_row) else None
                if source_value != gold_value:
                    changed[header] = gold_value
            if changed:
                expected_changes += len(changed)
                updates.append((sheet_name, row_number, changed))
    source.close()
    gold.close()
    assert expected_changes > 100_000

    destination = TEST_ROOT / "full-acceptance.xlsx"
    export_reviewed_workbook(
        EMPTY_WORKBOOK,
        destination,
        updates,
        column_definitions=column_definitions,
        sheet_dimensions=sheet_dimensions,
    )
    actual = load_workbook(destination, read_only=True, data_only=False)
    expected = load_workbook(GOLD_WORKBOOK, read_only=True, data_only=False)
    mismatches = defaultdict(int)
    for sheet_name in expected.sheetnames:
        actual_ws = actual[sheet_name]
        expected_ws = expected[sheet_name]
        assert actual_ws.max_row == expected_ws.max_row
        assert actual_ws.max_column == expected_ws.max_column
        for actual_row, expected_row in zip(
            actual_ws.iter_rows(values_only=True),
            expected_ws.iter_rows(values_only=True),
            strict=False,
        ):
            if tuple(map(normalize_blank, actual_row)) != tuple(map(normalize_blank, expected_row)):
                mismatches[sheet_name] += 1
    actual.close()
    expected.close()
    assert not mismatches
